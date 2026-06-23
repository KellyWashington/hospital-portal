from datetime import datetime, timedelta, time
import pandas as pd
import numpy as np
import re
from difflib import get_close_matches
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import os

# =====================================
# PATHS
# =====================================

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
INPUT_DIR  = os.path.join(BASE_DIR, "input")
OUTPUT_DIR = os.path.join(BASE_DIR, "output")

attendance_path      = os.path.join(INPUT_DIR,  "attendance.xls")
shifts_path          = os.path.join(INPUT_DIR,  "shifts.xlsx")
shifts_original_path = os.path.join(INPUT_DIR,  "shifts_original.xlsx")  # old file with Shifts_Config sheet
output_path          = os.path.join(OUTPUT_DIR, "Attendance_Report.xlsx")

os.makedirs(OUTPUT_DIR, exist_ok=True)

# =====================================
# LOAD FILES
# =====================================

print("Loading attendance file...")
try:
    attendance = pd.read_excel(attendance_path, engine="xlrd")
except Exception as e:
    print(f"Error loading attendance.xls: {e}")
    exit(1)

# Load shifts.xlsx (with Rest_Day column). Fall back to shifts_original.xlsx only if
# shifts.xlsx is missing entirely — but shifts_original.xlsx has no Rest_Day column.
shifts_path_to_read = shifts_path
if not os.path.exists(shifts_path_to_read):
    if os.path.exists(os.path.join(INPUT_DIR, "shifts_original.xlsx")):
        shifts_path_to_read = os.path.join(INPUT_DIR, "shifts_original.xlsx")
        print("[WARNING] shifts.xlsx not found; falling back to shifts_original.xlsx (no Rest_Day column — weekend half-day rule disabled)")

print(f"Loading shifts from {shifts_path_to_read}...")
try:
    # Auto-detect header row: some versions of shifts.xlsx have a merged note in row 0,
    # so the real column headers ("Name", "Department", ...) sit in row 1 (header=1).
    shifts_df = pd.read_excel(shifts_path_to_read)
    _tmp_cols = [str(c).strip() for c in shifts_df.columns]
    if "Name" not in _tmp_cols:
        shifts_df = pd.read_excel(shifts_path_to_read, header=1)
except Exception as e:
    print(f"Error loading shifts file: {e}")
    exit(1)

# ── Load Shifts_Config from shifts_original.xlsx (the old file with a Shifts_Config sheet).
# This table drives get_closest_shift for ALL shift groups — FO rotating, Lab, Nurse/PA,
# Janitor, Default.  The script falls back to hardcoded values only if the sheet is absent.
shifts_config_lookup: dict[str, dict] = {}   # Shift_ID → {Start_Time, End_Time, Is_Night_Shift}
_config_src = shifts_original_path if os.path.exists(shifts_original_path) else None
# Also accept shifts.xlsx itself when it contains a Shifts_Config sheet (old-format file)
if _config_src is None and os.path.exists(shifts_path):
    try:
        _xl = pd.ExcelFile(shifts_path)
        if "Shifts_Config" in _xl.sheet_names:
            _config_src = shifts_path
    except Exception:
        pass

if _config_src:
    try:
        _cfg = pd.read_excel(_config_src, sheet_name="Shifts_Config")
        _cfg.columns = [str(c).strip() for c in _cfg.columns]
        for _, r in _cfg.iterrows():
            sid = str(r["Shift_ID"]).strip()
            shifts_config_lookup[sid] = {
                "Shift_ID":       sid,
                "Start_Time":     str(r["Start_Time"]).strip(),
                "End_Time":       str(r["End_Time"]).strip(),
                "Is_Night_Shift": str(r["Is_Night_Shift"]).strip().lower() in ("yes", "true", "1"),
            }
        print(f"Loaded {len(shifts_config_lookup)} shift configs from {_config_src}")
    except Exception as e:
        print(f"[WARNING] Could not load Shifts_Config sheet: {e} — using hardcoded fallbacks")
else:
    print("[INFO] No Shifts_Config sheet found; all shift windows are hardcoded in the script")

def _cfg(shift_id: str, start: str, end: str, night: bool) -> dict:
    """Return the Shifts_Config row for shift_id, falling back to the provided hardcoded values."""
    if shift_id in shifts_config_lookup:
        return shifts_config_lookup[shift_id]
    return {"Shift_ID": shift_id, "Start_Time": start, "End_Time": end, "Is_Night_Shift": night}

# Clean columns
attendance.columns = [str(c).strip() for c in attendance.columns]
shifts_df.columns = [str(c).strip() for c in shifts_df.columns]

# Parse timestamps
attendance["Time"]  = pd.to_datetime(attendance["Time"], dayfirst=True, errors="coerce")
attendance["State"] = attendance["State"].astype(str).str.strip()

# Keep only standard punch events
attendance = attendance[attendance["State"].isin(["C/In", "C/Out"])].copy()

# =====================================
# NAME NORMALIZATION
# =====================================

def normalize_name(name):
    return re.sub(r"[^a-z0-9 ]", "", str(name).lower()).strip()

attendance["norm_name"] = attendance["Name"].apply(normalize_name)
shifts_df["norm_name"] = shifts_df["Name"].apply(normalize_name)

EXCLUDED_EMPLOYEES = {
    "wickliff ondiba",  # assigned to sister facility; exclude from this site's report
}
if EXCLUDED_EMPLOYEES:
    before_excluded = len(attendance)
    attendance = attendance[~attendance["norm_name"].isin(EXCLUDED_EMPLOYEES)].copy()
    removed_excluded = before_excluded - len(attendance)
    if removed_excluded:
        print(f"[INFO] Excluded {removed_excluded} punch(es) for employees assigned away from this facility")

# Map names to their shift strings
# Column may be named 'Override_Shift' (current format), 'Shift', or 'SHIFT' (legacy)
for _candidate in ("Override_Shift", "Shift", "SHIFT"):
    if _candidate in shifts_df.columns:
        shift_col = _candidate
        break
else:
    # Last-resort: pick any column that isn't Name/Department/norm_name
    _non_key = [c for c in shifts_df.columns if c not in ("Name", "Department", "norm_name")]
    shift_col = _non_key[0] if _non_key else "Override_Shift"
    print(f"[WARNING] Could not find a shift column; using '{shift_col}' as fallback")

shift_map = dict(zip(shifts_df["norm_name"], shifts_df[shift_col].fillna("")))

# Map names to department
dept_col = "Department" if "Department" in shifts_df.columns else None
dept_map = dict(zip(shifts_df["norm_name"], shifts_df[dept_col].fillna("") if dept_col else ""))

# ── Load Department_Rules sheet → maps department label to shift pool IDs ─────
dept_rules: dict[str, list[str]] = {}
try:
    _rules_df = pd.read_excel(shifts_path_to_read, sheet_name="Department_Rules")
    _rules_df.columns = [str(c).strip() for c in _rules_df.columns]
    for _, r in _rules_df.iterrows():
        dept_label = str(r["Department"]).strip()
        shift_ids  = [s.strip() for s in str(r["Allowed_Shifts"]).split(",") if s.strip()]
        dept_rules[dept_label] = shift_ids
    print(f"Loaded department rules for: {list(dept_rules.keys())}")
except Exception as e:
    print(f"[WARNING] Could not load Department_Rules sheet: {e} — using hardcoded fallbacks")

# ── Warn about employees in attendance but not in shifts file ─────────────────
import re as _re
_att_norms   = set(attendance["norm_name"].unique())
_shift_norms = set(shifts_df["norm_name"].unique())
_unmatched   = sorted(_att_norms - _shift_norms)
if _unmatched:
    print(f"[WARNING] {len(_unmatched)} employee(s) in attendance have no shift config — defaulting to 08:00-17:00:")
    for _u in _unmatched:
        print(f"  · '{_u}'")

# Map names to their designated rest day: "Sunday" or "Saturday" (standard employees only)
# "Sunday"  → employee works Mon–Sat; Sunday punches are ignored
# "Saturday" → employee works Sun–Fri; Saturday punches are ignored
# Blank / NaN → no rest-day rule (Nurse/PA, Laboratory, night staff)
if "Rest_Day" in shifts_df.columns:
    rest_day_map = dict(zip(shifts_df["norm_name"], shifts_df["Rest_Day"].fillna("")))
else:
    print("[INFO] No 'Rest_Day' column found in shifts file — weekend half-day rule disabled")
    rest_day_map = {}

# =====================================
# INTERNAL ALLOWED SHIFTS DYNAMIC PARSER
# =====================================

def get_allowed_shifts_for_employee(norm_name, emp_name, session_date=None):
    """
    Returns the allowed shift configs for an employee on a given date.

    Shift detection priority:
      1. Explicit Override_Shift text (6pm-8am, 11am-8pm, FO_ROTATING, etc.)
      2. Department_Rules sheet (loaded from shifts.xlsx) — uses exact dept label from file
      3. Legacy hardcoded fallbacks for FRONT OFFICE and any unrecognised dept
      4. Final fallback: 8am-5pm

    Weekend half-day rule (standard day-shift employees ONLY):
      Controlled by Rest_Day column in shifts.xlsx.
      Night-shift workers are EXEMPT — Saturday half-day is never applied when
      the assigned shift is a night shift.
    """
    shift_text = shift_map.get(norm_name, "")
    shift_text = "" if pd.isna(shift_text) else str(shift_text).strip()

    # Keep raw dept label (exact case from file) for dept_rules lookup
    department_raw = str(dept_map.get(norm_name, "")).strip()
    department     = department_raw.upper()   # for legacy uppercase checks
    rest_day       = str(rest_day_map.get(norm_name, "")).strip().title()

    txt = shift_text.lower().replace(" ", "")

    # ── Identify if this employee has a dedicated night shift ──────────────────
    is_dedicated_night = any(x in txt for x in ["6pm-8", "6pm-7am", "6pm-7:30", "6:30pm-8"])

    # ── Departments that may work nights (exempt from weekend half-day) ────────
    NIGHT_EXEMPT_DEPT_LABELS = {
        "NURSES", "NURSE/PA", "HOUSEKEEPING( PA)", "HOUSEKEEPING(JANITORS)",
        "LABORATORY", "THEATRE", "FRONT OFFICE",
        "Nurse/PA", "Laboratory",   # exact values from current shifts.xlsx
    }
    is_exempt_dept = (department in NIGHT_EXEMPT_DEPT_LABELS or
                      department_raw in NIGHT_EXEMPT_DEPT_LABELS)

    # ── Standard employee with a Rest_Day defined ─────────────────────────────
    is_standard = (
        rest_day in ("Sunday", "Saturday") and
        not is_dedicated_night and
        not is_exempt_dept
    )

    if is_standard and session_date is not None:
        day_name = pd.Timestamp(session_date).day_name()
        if day_name == rest_day:
            return []   # Rest day — caller marks session as "Not Scheduled"
        if day_name in ("Saturday", "Sunday"):
            return [{"Shift_ID": "Weekend_Half", "Start_Time": "08:00", "End_Time": "14:00", "Is_Night_Shift": False}]

    # ── 1. Explicit shift text overrides (always take priority) ───────────────

    if "6:30pm" in txt or "630pm" in txt:
        return [{"Shift_ID": "FO_Night_630", "Start_Time": "18:30", "End_Time": "08:00", "Is_Night_Shift": True}]

    if "6pm-8" in txt or "6pm- 8" in txt:
        return [{"Shift_ID": "Night_6to8", "Start_Time": "18:00", "End_Time": "08:00", "Is_Night_Shift": True}]

    if "11am-8" in txt or "11am8pm" in txt:
        return [{"Shift_ID": "Day_11to8", "Start_Time": "11:00", "End_Time": "20:00", "Is_Night_Shift": False}]

    if "10am-7" in txt or "10am7pm" in txt or "10:00-19:00" in txt:
        return [{"Shift_ID": "Day_10to7", "Start_Time": "10:00", "End_Time": "19:00", "Is_Night_Shift": False}]

    if "9am-6" in txt or "9am6pm" in txt or "09:00-18:00" in txt or "9:00-18:00" in txt:
        return [{"Shift_ID": "Day_9to6", "Start_Time": "09:00", "End_Time": "18:00", "Is_Night_Shift": False}]

    if shift_text == "ORTHO_10_ROTATING":
        return [
            _cfg("Default_Day", "08:00", "17:00", False),
            _cfg("Day_10to7",  "10:00", "19:00", False),
        ]

    if shift_text == "ORTHO_11_ROTATING":
        return [
            _cfg("Default_Day", "08:00", "17:00", False),
            _cfg("Day_11to8",  "11:00", "20:00", False),
        ]

    if shift_text == "THEATRE_ROTATING":
        return [
            _cfg("PA_Nurse_Day", "07:30", "18:00", False),
            _cfg("Day_10to7",    "10:00", "19:00", False),
            _cfg("Day_11to8",    "11:00", "20:00", False),
        ]

    if department in ("OTHER", "DEFAULT", ""):
        return [
            _cfg("Default_Day", "08:00", "17:00", False),
            _cfg("Day_10to7",  "10:00", "19:00", False),
            _cfg("Day_11to8",  "11:00", "20:00", False),
        ]

    if shift_text == "FO_ROTATING":
        return [
            _cfg("FO_Day_8",  "08:00", "17:00", False),
            _cfg("FO_Day_9",  "09:00", "18:00", False),
            _cfg("FO_Day_10", "10:00", "19:00", False),
            _cfg("FO_Day_11", "11:00", "20:00", False),
            _cfg("FO_Night",  "18:30", "08:00", True),
        ]

    # ── 2. Department_Rules sheet → shift pool lookup ─────────────────────────
    # SHIFT_ID_DEFAULTS provides start/end/night for every known Shift_ID so the
    # script works correctly even when Shifts_Config sheet is absent.
    SHIFT_ID_DEFAULTS = {
        "Default_Day":    ("08:00", "17:00", False),
        "PA_Nurse_Day":   ("07:30", "18:00", False),
        "PA_Nurse_Night": ("18:00", "07:30", True),
        "Lab_Day_7":      ("07:00", "16:00", False),
        "Lab_Day_8":      ("08:00", "17:00", False),
        "Lab_Day_9":      ("09:00", "18:00", False),
        "Lab_Night":      ("18:00", "07:00", True),
        "Janitor_Day":    ("07:00", "16:00", False),
        "Janitor_Night":  ("18:00", "07:00", True),
        "FO_Day_8":       ("08:00", "17:00", False),
        "FO_Day_9":       ("09:00", "18:00", False),
        "FO_Day_10":      ("10:00", "19:00", False),
        "FO_Day_11":      ("11:00", "20:00", False),
        "FO_Night":       ("18:30", "08:00", True),
        "Night_6to8":     ("18:00", "08:00", True),
        "FO_Night_630":   ("18:30", "08:00", True),
        "Day_11to8":      ("11:00", "20:00", False),
        "Day_10to7":      ("10:00", "19:00", False),
        "Day_9to6":       ("09:00", "18:00", False),
        "Weekend_Half":   ("08:00", "14:00", False),
    }

    # Builtin department → shift pool (used when Department_Rules sheet is absent or
    # the employee's department isn't listed there). This guarantees nurses, lab, etc.
    # NEVER fall through to Default_Day.
    BUILTIN_DEPT_RULES = {
        "NURSES":                  ["PA_Nurse_Day", "PA_Nurse_Night"],
        "NURSE":                   ["PA_Nurse_Day", "PA_Nurse_Night"],
        "NURSE/PA":                ["PA_Nurse_Day", "PA_Nurse_Night"],
        "HOUSEKEEPING( PA)":       ["PA_Nurse_Day", "PA_Nurse_Night"],
        "HOUSEKEEPING(JANITORS)":  ["Janitor_Day",  "Janitor_Night"],
        "HOUSEKEEPING(JANITORS) ": ["Janitor_Day",  "Janitor_Night"],
        "LABORATORY":              ["Lab_Day_7", "Lab_Day_8", "Lab_Day_9", "Lab_Night"],
        "THEATRE":                 ["PA_Nurse_Day", "PA_Nurse_Night"],
        "CLINICAL":                ["Default_Day", "FO_Day_9", "Day_11to8", "Night_6to8"],
        "FRONT OFFICE":            ["FO_Day_8", "FO_Day_9", "FO_Day_10", "FO_Day_11", "FO_Night"],
        "ORTHOPEDIC":              ["Default_Day", "Day_10to7", "Day_11to8"],
        "ORTHOPAEDIC":             ["Default_Day", "Day_10to7", "Day_11to8"],
    }

    def _build_pool(shift_ids):
        pool = []
        for sid in shift_ids:
            if sid in SHIFT_ID_DEFAULTS:
                s, e, n = SHIFT_ID_DEFAULTS[sid]
                pool.append(_cfg(sid, s, e, n))
            else:
                pool.append(_cfg(sid, "08:00", "17:00", False))
        return pool

    # Try Department_Rules sheet first (exact match on raw dept label from file)
    if department_raw in dept_rules:
        pool = _build_pool(dept_rules[department_raw])
        if pool:
            return pool

    # Builtin rules (uppercase dept match — catches any casing variant in the file)
    if department in BUILTIN_DEPT_RULES:
        return _build_pool(BUILTIN_DEPT_RULES[department])

    # Also try stripping trailing spaces and normalising slash spacing
    dept_normalised = department.replace(" ", "").replace("/", "/")
    for key in BUILTIN_DEPT_RULES:
        if key.replace(" ", "").replace("/", "/") == dept_normalised:
            return _build_pool(BUILTIN_DEPT_RULES[key])

    # ── 3. Final fallback: 8am–5pm ────────────────────────────────────────────
    return [_cfg("Default_Day", "08:00", "17:00", False)]

def parse_time_str(t_str):
    h, m = map(int, t_str.split(":"))
    return h, m

# =====================================
# SHIFT MATCHING ENGINE
# =====================================

def get_closest_shift(ref_time, allowed_configs, is_checkout_only=False):
    """
    Finds the shift whose Start_Time (or End_Time if checkout-only) is closest to ref_time.

    Tiebreaker for arrivals:  when a C/In time is equidistant between a day-shift END
    and a night-shift START (e.g. 17:30 is 30 min from both Nurse_Day end=18:00 and
    Nurse_Night start=18:00), prefer the night shift — an arrival near 17:30-18:30 is
    almost always the start of an overnight shift, not a late-day re-arrival.

    For checkouts the mirror logic applies: a C/Out in the range 06:00-09:00 is almost
    always the end of a night shift, so prefer night-shift end over day-shift start.
    """
    best_config = None
    best_diff = None

    for sh in allowed_configs:
        sh_h, sh_m = parse_time_str(sh["End_Time"] if is_checkout_only else sh["Start_Time"])

        # Same day
        dt = datetime.combine(ref_time.date(), time(sh_h, sh_m))
        diff = abs((ref_time - dt).total_seconds())

        # Previous and next days for overnight shifts
        if sh["Is_Night_Shift"]:
            dt_prev = datetime.combine(ref_time.date() - timedelta(days=1), time(sh_h, sh_m))
            diff_prev = abs((ref_time - dt_prev).total_seconds())
            if diff_prev < diff:
                diff = diff_prev

            dt_next = datetime.combine(ref_time.date() + timedelta(days=1), time(sh_h, sh_m))
            diff_next = abs((ref_time - dt_next).total_seconds())
            if diff_next < diff:
                diff = diff_next

        # Tiebreaker: on exact tie (or within 1 second), prefer night shift for evening
        # arrivals (17:00–23:59) and for morning departures (00:00–10:00)
        if best_diff is not None and abs(diff - best_diff) <= 1:
            ref_hour = ref_time.hour
            if not is_checkout_only:
                # Arrival: prefer night shift when checking in during evening hours
                prefer_night = (17 <= ref_hour <= 23)
            else:
                # Departure: prefer night shift when checking out in early morning
                prefer_night = (0 <= ref_hour <= 10)
            if prefer_night and sh["Is_Night_Shift"] and not best_config["Is_Night_Shift"]:
                best_diff = diff
                best_config = sh
                continue

        if best_diff is None or diff < best_diff:
            best_diff = diff
            best_config = sh

    return best_config

# =====================================
# LATENESS AND GRACE CALCULATION
# =====================================

GRACE_MINUTES = 6

def calculate_lateness_mins(cin_time, shift_config):
    sh_h, sh_m = parse_time_str(shift_config["Start_Time"])
    shift_start_dt = datetime.combine(cin_time.date(), time(sh_h, sh_m))

    if shift_config["Is_Night_Shift"]:
        if cin_time.hour < 12:
            # Early-morning C/In = night worker finishing from yesterday's shift start
            # Adjust to previous day's shift start for lateness comparison
            shift_start_dt = datetime.combine(cin_time.date() - timedelta(days=1), time(sh_h, sh_m))
        elif cin_time.hour < sh_h:
            # C/In is before shift start on same day (e.g. 17:33 for 18:00 shift) — early arrival
            # Keep same-day shift_start_dt; diff will be negative → late_mins = 0 (early)
            pass
        # else: C/In is after shift start on same day → normal lateness calc

    diff_mins = int((cin_time - shift_start_dt).total_seconds() / 60)

    # Early arrivals must NEVER count as late
    late_mins = max(0, diff_mins - GRACE_MINUTES)

    if late_mins == 0:
        status = "On Time"
    elif late_mins <= 15:
        status = "Slightly Late"
    elif late_mins <= 60:
        status = "Late"
    else:
        status = "Very Late"

    return late_mins, status

# =====================================
# RAW PUNCH DEDUPLICATION
# =====================================

# Deduplicate on employee + time + state (exact same biometric double-tap)
attendance = attendance.drop_duplicates(subset=["norm_name", "Time", "State"])

# =====================================
# PUNCH INTENT CLASSIFICATION
# =====================================
# The biometric terminal does NOT reset between employees — it inherits the previous
# person's state.  A day-shift employee arriving at 08:00 may tap C/Out simply because
# the last user tapped C/In.  The same misfire happens at end-of-day (C/In instead of
# C/Out).  However, for NIGHT-CAPABLE departments (Nurses, Lab, PA, Janitors, FO,
# Theatre, Clinical, Radiology, Pharmacy) an evening C/In or a morning C/Out is
# perfectly real — it is a genuine night-shift start or finish — so for those employees
# we TRUST the State label as recorded.
#
# Classification rules
# ─────────────────────
# Night-capable employee  →  use State AS-IS
# Day-only employee       →  use TIME:
#       punch before 13:00  →  treat as ARRIVAL  (regardless of recorded State)
#       punch 13:00+        →  treat as DEPARTURE (regardless of recorded State)
#
# Whenever the recorded State differs from the time-derived intent, we set a
# correction flag so it appears in the Data Quality Flag column.

NIGHT_CAPABLE_DEPTS = {
    "NURSES", "NURSE", "NURSE/PA", "PA", "HOUSEKEEPING( PA)", "HOUSEKEEPING(JANITORS)",
    "LABORATORY", "FRONT OFFICE", "THEATRE",
    "CLINICAL", "RADIOLOGY", "PHARMACY",
}

def employee_is_night_capable(norm_name):
    dept = str(dept_map.get(norm_name, "")).strip().upper()
    shift_txt = str(shift_map.get(norm_name, "")).strip().lower()
    if dept in NIGHT_CAPABLE_DEPTS:
        return True
    # Only flag as night-capable when the shift genuinely starts in the evening.
    # "11am-8pm" contains "pm" but is a late-day shift, NOT a night shift —
    # so we check for 6pm or later start, not just any "pm" occurrence.
    shift_txt_compact = shift_txt.replace(" ", "")
    if (
        "night" in shift_txt_compact or
        shift_txt_compact.startswith(("6pm-", "18:00-", "18:30-", "6:30pm-", "630pm-"))
    ):
        return True

    # ── Fallback: infer from actual punch pattern ─────────────────────────────
    # Employees not found in the shifts file (unknown department) who consistently
    # check IN during evening hours (17:00–23:00) and check OUT in the morning
    # (00:00–12:00) are almost certainly night-shift workers.  Without this check
    # they default to day-only mode and have all their punches misclassified.
    if norm_name not in _shift_norms:
        emp_punches = attendance[attendance["norm_name"] == norm_name]
        evening_cins = emp_punches[
            (emp_punches["State"] == "C/In") &
            (emp_punches["Time"].dt.hour >= 17) &
            (emp_punches["Time"].dt.hour <= 23)
        ]
        morning_couts = emp_punches[
            (emp_punches["State"] == "C/Out") &
            (emp_punches["Time"].dt.hour <= 10)
        ]
        total_cins = (emp_punches["State"] == "C/In").sum()
        # If majority of C/Ins are in the evening → night worker
        if total_cins > 0 and len(evening_cins) / total_cins >= 0.5:
            return True
        # Or if they have morning C/Outs with no matching morning C/Ins
        if len(morning_couts) >= 2:
            return True
    # ─────────────────────────────────────────────────────────────────────────

    return False

def classify_punch(row, is_night_capable):
    """
    Returns (intent, corrected_flag, correction_note)
    intent: 'arrival' | 'departure'
    """
    h      = row["Time"].hour
    state  = row["State"]
    t_str  = row["Time"].strftime("%H:%M")

    if is_night_capable:
        # Trust the recorded State label
        if state == "C/In":
            return "arrival",   False, ""
        else:
            return "departure", False, ""
    else:
        # Day-only: reclassify by time
        if h < 13:
            corrected = (state != "C/In")
            note = f"arrival corrected (was C/Out at {t_str})" if corrected else ""
            return "arrival", corrected, note
        else:
            corrected = (state != "C/Out")
            note = f"departure corrected (was C/In at {t_str})" if corrected else ""
            return "departure", corrected, note

# =====================================
# SESSION BUILDING
# =====================================
# A session = one work period = one (arrival, departure) pair.
# Algorithm (per employee, chronologically):
#
#   1. Collapse duplicate taps (same-second exact duplicates already removed above;
#      here we collapse same-intent punches within DOUBLE_TAP_WINDOW to single best).
#   2. Pair: for each unpaired arrival find the NEAREST following departure within
#      PAIR_WINDOW.  For night-capable employees a departure that crosses midnight
#      is still within PAIR_WINDOW.
#   3. Unpaired arrivals → Missing C/Out.
#   4. Unpaired departures → Missing C/In.
#
# Days Attended:  every session (row) counts — even if only one side exists.
# Lateness:       calculated whenever an arrival exists.
# Missing flags:  set individually per session, do NOT suppress lateness.

sessions = []
PAIR_WINDOW       = timedelta(hours=16)
DOUBLE_TAP_WINDOW = timedelta(minutes=5)

print("Grouping punches into sessions using Intent-Aware Pairing...")

for emp, grp in attendance.groupby("norm_name"):
    grp      = grp.sort_values("Time").reset_index(drop=True)
    emp_name = grp["Name"].iloc[0]
    night_cap = employee_is_night_capable(emp)

    # ── Step 1: classify intent for every punch ──────────────────────────────
    classified = []
    for _, row in grp.iterrows():
        intent, corrected, note = classify_punch(row, night_cap)
        classified.append({
            "Time":      row["Time"],
            "State":     row["State"],
            "intent":    intent,
            "corrected": corrected,
            "note":      note,
        })

    # ── Step 2a: discard same-second C/In when a C/Out exists at the same time ─
    # The terminal occasionally records a single tap as both C/In and C/Out at the
    # exact same second.  The C/Out is the correct interpretation (the employee is
    # leaving); discard the phantom C/In so it doesn't seed a second session.
    same_second_out_times = set()
    for p in classified:
        if p["intent"] == "departure":
            same_second_out_times.add(p["Time"])
    classified = [
        p for p in classified
        if not (p["intent"] == "arrival" and p["Time"] in same_second_out_times)
    ]

    # ── Step 2b: collapse double-taps within DOUBLE_TAP_WINDOW ───────────────
    # Same intent within 5 min: keep earliest arrival, latest departure.
    # Mixed intent within 5 min: keep both (e.g. two people tapped back-to-back).
    deduped  = []
    used     = set()
    n        = len(classified)

    for i in range(n):
        if i in used:
            continue
        used.add(i)
        cluster = [classified[i]]
        for j in range(i + 1, n):
            if j in used:
                continue
            if (classified[j]["Time"] - cluster[-1]["Time"]) <= DOUBLE_TAP_WINDOW:
                cluster.append(classified[j])
                used.add(j)
            else:
                break

        arrivals   = [p for p in cluster if p["intent"] == "arrival"]
        departures = [p for p in cluster if p["intent"] == "departure"]
        # Keep best of each kind present in the cluster
        if arrivals:
            deduped.append(min(arrivals,   key=lambda p: p["Time"]))
        if departures:
            deduped.append(max(departures, key=lambda p: p["Time"]))

    deduped.sort(key=lambda p: p["Time"])

    # ── Step 2c: collapse same-day all-arrival sequences (no departures) ──────
    # If an employee has 2+ arrivals in one calendar day with zero departures,
    # the later tap(s) are stray re-taps, NOT new shift starts.
    # Keep only the first arrival for that day; drop the rest.
    #
    # For night-capable employees the same rule applies with one extra condition:
    # only collapse when the first arrival has NO matching departure at all —
    # if there is a departure (even the next morning), both sessions are real.
    from collections import defaultdict
    by_date = defaultdict(list)
    for p in deduped:
        by_date[p["Time"].date()].append(p)
    deduped_clean = []
    for day_punches in sorted(by_date.keys()):
        punches = by_date[day_punches]
        arrivals_today   = [p for p in punches if p["intent"] == "arrival"]
        departures_today = [p for p in punches if p["intent"] == "departure"]
        if len(arrivals_today) > 1 and len(departures_today) == 0:
            # All arrivals, no departures — keep first arrival only
            first = min(arrivals_today, key=lambda p: p["Time"])
            stray_times = ", ".join(
                p["Time"].strftime("%H:%M") for p in arrivals_today if p is not first
            )
            first = dict(first)
            existing_note = first.get("note", "")
            first["note"] = (existing_note + f"; stray re-tap(s) discarded: {stray_times}").lstrip("; ")
            deduped_clean.append(first)
        else:
            deduped_clean.extend(punches)
    deduped_clean.sort(key=lambda p: p["Time"])
    deduped = deduped_clean

    # ── Step 2d: collapse consecutive same-intent orphan departures (day-only) ─
    # For day-only employees, multiple departure-intent punches in a row collapse to
    # a single orphan departure per calendar day (the latest one on that day).
    # If the first departure in the run has a preceding arrival, we keep it to pair
    # with the arrival, and keep the latest departure of any subsequent days.
    if not night_cap:
        collapsed = []
        i = 0
        while i < len(deduped):
            curr = deduped[i]
            if curr["intent"] == "departure":
                run = [curr]
                j = i + 1
                while j < len(deduped) and deduped[j]["intent"] == "departure":
                    run.append(deduped[j])
                    j += 1
                
                # Group departures in this run by calendar date and keep the latest for each date
                from collections import defaultdict as _defaultdict
                by_date_run = _defaultdict(list)
                for p in run:
                    by_date_run[p["Time"].date()].append(p)
                
                # For each date, keep only the latest departure
                collapsed_run = []
                for d in sorted(by_date_run.keys()):
                    collapsed_run.append(max(by_date_run[d], key=lambda p: p["Time"]))
                
                # Check if the first departure in this run has a preceding arrival.
                # A preceding arrival exists if the last item in 'collapsed' is an arrival.
                has_preceding_arrival = False
                if collapsed and collapsed[-1]["intent"] == "arrival":
                    has_preceding_arrival = True
                
                if has_preceding_arrival:
                    # Keep the first departure of the run to pair with the arrival
                    collapsed.append(collapsed_run[0])
                    # If there are more departures in the run (on subsequent days),
                    # they are orphans; keep them.
                    if len(collapsed_run) > 1:
                        collapsed.extend(collapsed_run[1:])
                else:
                    # No preceding arrival: all departures in this run are orphans.
                    # Keep the latest departure of each day.
                    collapsed.extend(collapsed_run)
                i = j
            else:
                collapsed.append(curr)
                i += 1
        deduped = collapsed

    n_d = len(deduped)

    # ── Step 3: pair arrivals to departures ──────────────────────────────────
    used_pair       = set()
    paired_sessions = []

    for i in range(n_d):
        if i in used_pair:
            continue
        curr = deduped[i]
        if curr["intent"] != "arrival":
            continue

        best_j = None
        for j in range(i + 1, n_d):
            if j in used_pair:
                continue
            nxt  = deduped[j]
            diff = nxt["Time"] - curr["Time"]
            if diff > PAIR_WINDOW:
                break
            if nxt["intent"] == "departure":
                best_j = j
                break   # greedy: first departure wins

        flags = []
        if curr["corrected"]:
            flags.append(curr["note"])

        if best_j is not None:
            dep = deduped[best_j]
            if dep["corrected"]:
                flags.append(dep["note"])
            used_pair.add(i)
            used_pair.add(best_j)
            paired_sessions.append({
                "Check-In":  curr["Time"],
                "Check-Out": dep["Time"],
                "Flags":     "; ".join(flags),
            })
        else:
            used_pair.add(i)
            flags.append("Missing C/Out")
            paired_sessions.append({
                "Check-In":  curr["Time"],
                "Check-Out": pd.NaT,
                "Flags":     "; ".join(flags),
            })

    # ── Step 4: orphan departures (no preceding arrival) ─────────────────────
    for i in range(n_d):
        if i in used_pair:
            continue
        curr  = deduped[i]   # must be a departure at this point
        flags = []
        if curr["corrected"]:
            flags.append(curr["note"])
        flags.append("Missing C/In")
        paired_sessions.append({
            "Check-In":  pd.NaT,
            "Check-Out": curr["Time"],
            "Flags":     "; ".join(flags),
        })

    # ── Step 5: discard phantom arrivals that fall inside a completed pair ──────
    # If an arrival lands between a paired session's C/In and C/Out, it is a stray
    # tap during an active shift (e.g. someone re-tapped mid-day).  Remove those
    # orphan arrivals so they don't become spurious Incomplete sessions.
    paired_intervals = [
        (s["Check-In"], s["Check-Out"])
        for s in paired_sessions
        if not pd.isna(s.get("Check-In")) and not pd.isna(s.get("Check-Out"))
    ]
    cleaned_sessions = []
    for s in paired_sessions:
        if pd.isna(s.get("Check-Out")) and not pd.isna(s.get("Check-In")):
            cin_time = s["Check-In"]
            is_phantom = any(
                lo < cin_time < hi
                for lo, hi in paired_intervals
                if lo != cin_time
            )
            if is_phantom:
                continue  # drop this orphan arrival — it's inside another session
        cleaned_sessions.append(s)
    paired_sessions = cleaned_sessions

    # ── Step 6: sort sessions chronologically ────────────────────────────────
    paired_sessions.sort(key=lambda s: s["Check-In"] if not pd.isna(s["Check-In"]) else s["Check-Out"])
            
    # ── Step 6: classify each session ───────────────────────────────────────
    # Get weekday allowed shifts for this employee (date-agnostic; weekend
    # override is applied per session below)
    allowed_configs_weekday = get_allowed_shifts_for_employee(emp, emp_name, session_date=None)

    for s in paired_sessions:
        cin  = s["Check-In"]
        cout = s["Check-Out"]

        # ── Determine session date ────────────────────────────────────────────
        # For night-capable employees a session that starts ~18:00-23:59 and ends
        # the next morning belongs to the EVENING date (start date).
        # For a Missing-C/In session where the only punch is a morning departure,
        # it may belong to the previous calendar day (overnight carry-over).
        if not pd.isna(cin):
            session_date = cin.date()
        else:
            # Only C/Out recorded
            session_date = cout.date()
            if night_cap and cout.hour < 13:
                # Could be the tail of a night shift that started yesterday
                temp = get_closest_shift(cout, allowed_configs_weekday, is_checkout_only=True)
                if temp and temp["Is_Night_Shift"]:
                    session_date = (cout - timedelta(days=1)).date()

        # ── Get shift pool for this specific date (weekend half-day rule) ─────
        allowed_configs = get_allowed_shifts_for_employee(emp, emp_name, session_date=session_date)

        # ── Rest Day ──────────────────────────────────────────────────────────
        if len(allowed_configs) == 0:
            sessions.append({
                "Employee":         emp_name,
                "Date":             session_date,
                "Shift":            "Rest Day",
                "Check-In":         cin,
                "Check-Out":        cout,
                "Late Minutes":     0,
                "Attendance Status":"Rest Day — Not Scheduled",
                "Hours Worked":     np.nan,
                "Day Attendance":   "Rest Day",
                "Missing Checkin":  "Yes" if pd.isna(cin)  else "No",
                "Missing Checkout": "Yes" if pd.isna(cout) else "No",
                "Night Shift":      "No",
                "Data Quality Flag":s["Flags"] + (" | Punched on rest day" if s["Flags"] else "Punched on rest day"),
            })
            continue

        # ── Assign closest shift ──────────────────────────────────────────────
        if not pd.isna(cin):
            assigned = get_closest_shift(cin, allowed_configs, is_checkout_only=False)
        else:
            assigned = get_closest_shift(cout, allowed_configs, is_checkout_only=True)

        is_night_session = assigned["Is_Night_Shift"]
        shift_name       = f"{assigned['Start_Time']}-{assigned['End_Time']}"

        # ── Lateness — always from C/In when available ────────────────────────
        late_mins = 0
        if not pd.isna(cin):
            late_mins, lateness_raw = calculate_lateness_mins(cin, assigned)
            # Append missing-checkout note to status, do NOT suppress lateness info
            if pd.isna(cout):
                lateness_status = lateness_raw + " (Missing C/Out)"
            else:
                lateness_status = lateness_raw
        else:
            lateness_status = "Unknown — no C/In recorded"

        # ── Hours Worked ──────────────────────────────────────────────────────
        hours_worked = np.nan
        if not pd.isna(cin) and not pd.isna(cout):
            hours_worked = round((cout - cin).total_seconds() / 3600, 2)

        # ── Day Attendance category ───────────────────────────────────────────
        # Days Attended = every session that has at least one valid punch.
        # "Full Day"        — both C/In + C/Out present, checked out on time
        # "Early Departure" — both present, but left >90 min before shift end
        # "Incomplete"      — only one side recorded (missing C/In or C/Out)
        # "Anomalous"       — both present but punches fall outside shift window

        is_full_day  = False
        is_early_dep = False

        if not pd.isna(cin) and not pd.isna(cout):
            is_saturday = (pd.Timestamp(session_date).weekday() == 5)
            # Saturday half-day (08:00-14:00) only applies to day-shift workers.
            # Night-shift workers (Lab night, Nurse night, dedicated overnight) work
            # their normal shift on Saturdays — never override their window.
            if is_saturday and not is_night_session:
                eff_start_h, eff_start_m = 8,  0
                eff_end_h,   eff_end_m   = 14, 0
            else:
                eff_start_h, eff_start_m = parse_time_str(assigned["Start_Time"])
                eff_end_h,   eff_end_m   = parse_time_str(assigned["End_Time"])

            shift_start_dt = datetime.combine(session_date, time(eff_start_h, eff_start_m))
            shift_end_dt   = datetime.combine(session_date, time(eff_end_h,   eff_end_m))

            if is_night_session:
                # Night shift end is next calendar day
                shift_end_dt += timedelta(days=1)
                # Night shift start: if C/In is early morning it belongs to previous day's shift
                if cin.hour < 13:
                    shift_start_dt -= timedelta(days=1)

            cin_offset_mins  = (cin  - shift_start_dt).total_seconds() / 60
            cout_offset_mins = (cout - shift_end_dt).total_seconds()   / 60

            cin_in_window = -180 <= cin_offset_mins <= 180
            left_early    = cout_offset_mins < -90

            if cin_in_window and not left_early:
                is_full_day  = True
            elif cin_in_window and left_early:
                is_early_dep = True

        if is_full_day:
            day_attendance = "Full Day"
        elif is_early_dep:
            day_attendance = "Early Departure"
        elif pd.isna(cin) or pd.isna(cout):
            day_attendance = "Incomplete"
        else:
            day_attendance = "Anomalous"

        sessions.append({
            "Employee":         emp_name,
            "Date":             session_date,
            "Shift":            shift_name,
            "Check-In":         cin,
            "Check-Out":        cout,
            "Late Minutes":     late_mins,
            "Attendance Status":lateness_status,
            "Hours Worked":     hours_worked,
            "Day Attendance":   day_attendance,
            "Missing Checkin":  "Yes" if pd.isna(cin)  else "No",
            "Missing Checkout": "Yes" if pd.isna(cout) else "No",
            "Night Shift":      "Yes" if is_night_session else "No",
            "Data Quality Flag":s["Flags"],
        })

sessions_df = pd.DataFrame(sessions)
sessions_df["Data Quality Flag"] = sessions_df["Data Quality Flag"].fillna("")
sessions_df = sessions_df.sort_values(["Employee", "Date"]).reset_index(drop=True)

# Separate rest-day records from scheduled sessions for analytics
# (Rest Day rows are kept in sessions_df for full audit trail in the Sessions sheet)
scheduled_df = sessions_df[sessions_df["Day Attendance"] != "Rest Day"].copy()

# =====================================
# ANALYTICS & SUMMARIES (NO DEPARTMENTS)
# =====================================

print("Generating employee summaries...")

# ── Helper: extract base punctuality label from combined status ───────────
# Statuses look like "On Time", "Slightly Late (Missing C/Out)", "Unknown — no C/In recorded"
LATE_LABELS = {"On Time", "Slightly Late", "Late", "Very Late"}

def base_lateness(status_val):
    s = str(status_val)
    for lbl in LATE_LABELS:
        if s.startswith(lbl):
            return lbl
    return "Unknown"

scheduled_df["_base_status"] = scheduled_df["Attendance Status"].apply(base_lateness)

# ── Sessions where C/In was actually recorded (lateness denominator) ─────
cin_sessions = scheduled_df[scheduled_df["Missing Checkin"] == "No"]

# 1. Detailed Employee Summary Sheet (For HR)
employee_summary = scheduled_df.groupby("Employee").agg(
    Total_Sessions        = ("Date", "count"),
    # Days Attended = every session regardless of completeness
    Days_Attended         = ("Day Attendance", lambda x: x.isin(["Full Day","Early Departure","Incomplete","Anomalous"]).sum()),
    Full_Days             = ("Day Attendance", lambda x: (x == "Full Day").sum()),
    Early_Departures      = ("Day Attendance", lambda x: (x == "Early Departure").sum()),
    Incomplete_Days       = ("Day Attendance", lambda x: (x == "Incomplete").sum()),
    Total_Late_Minutes    = ("Late Minutes", "sum"),
    Average_Late_Minutes  = ("Late Minutes", "mean"),
    Average_Hours_Worked  = ("Hours Worked", "mean"),
    Total_Hours_Worked    = ("Hours Worked", "sum"),
    Missing_Checkin_Days  = ("Missing Checkin",  lambda x: (x == "Yes").sum()),
    Missing_Checkout_Days = ("Missing Checkout", lambda x: (x == "Yes").sum()),
    Flagged_Records       = ("Data Quality Flag", lambda x: (x != "").sum()),
).reset_index()

employee_summary = employee_summary[[
    "Employee", "Total_Sessions", "Days_Attended", "Full_Days", "Early_Departures", "Incomplete_Days",
    "Total_Late_Minutes", "Average_Late_Minutes", "Average_Hours_Worked", "Total_Hours_Worked",
    "Missing_Checkin_Days", "Missing_Checkout_Days", "Flagged_Records"
]]

# 2. Executive Status Summary Sheet
# Count base punctuality labels from C/In sessions only
status_counts = cin_sessions.groupby(["Employee", "_base_status"]).size().unstack(fill_value=0)
for cat in ["On Time", "Slightly Late", "Late", "Very Late"]:
    if cat not in status_counts.columns:
        status_counts[cat] = 0
status_counts = status_counts[[c for c in ["On Time","Slightly Late","Late","Very Late"] if c in status_counts.columns]]
status_counts = status_counts.reset_index()

# Overall session stats
summary_merg = scheduled_df.groupby("Employee").agg(
    Total_Sessions   = ("Date",            "count"),
    Days_Attended    = ("Day Attendance",  lambda x: x.isin(["Full Day","Early Departure","Incomplete","Anomalous"]).sum()),
    Full_Days        = ("Day Attendance",  lambda x: (x == "Full Day").sum()),
    Early_Departures = ("Day Attendance",  lambda x: (x == "Early Departure").sum()),
    Missing_Checkin  = ("Missing Checkin", lambda x: (x == "Yes").sum()),
    Missing_Checkout = ("Missing Checkout",lambda x: (x == "Yes").sum()),
).reset_index()

status_summary = pd.merge(summary_merg, status_counts, on="Employee", how="left").fillna(0)
for cat in ["On Time", "Slightly Late", "Late", "Very Late"]:
    status_summary[cat] = status_summary[cat].astype(int)

status_summary["Missing C/In (Unknown Lateness)"] = status_summary["Missing_Checkin"]

status_summary["Valid C/In Days"] = (
    status_summary["On Time"] + status_summary["Slightly Late"] +
    status_summary["Late"]    + status_summary["Very Late"]
)

status_summary["Late Days"] = (
    status_summary["Slightly Late"] + status_summary["Late"] + status_summary["Very Late"]
)

status_summary["Late Percentage"] = np.where(
    status_summary["Valid C/In Days"] > 0,
    ((status_summary["Late Days"] / status_summary["Valid C/In Days"]) * 100).round(1),
    0.0
)

status_summary = status_summary[[
    "Employee", "Days_Attended", "Full_Days", "Early_Departures",
    "On Time", "Slightly Late", "Late", "Very Late",
    "Missing C/In (Unknown Lateness)", "Missing_Checkout",
    "Total_Sessions", "Valid C/In Days", "Late Days", "Late Percentage"
]]

status_summary = status_summary.sort_values("Employee").reset_index(drop=True)

status_summary.rename(columns={
    "Days_Attended":   "Days Attended",
    "Full_Days":       "Full Days",
    "Early_Departures":"Early Departures",
    "Missing_Checkout":"Missing C/Out Days",
}, inplace=True)

# =====================================
# EXPORT EXCEL (WITH FILE LOCK PROTECTION)
# =====================================

print("Exporting workbook sheets...")
writer_path = output_path
try:
    with pd.ExcelWriter(writer_path, engine="openpyxl") as writer:
        status_summary.to_excel(writer,   sheet_name="Status Summary",    index=False)
        employee_summary.to_excel(writer, sheet_name="Employee Summary",  index=False)
        sessions_df.to_excel(writer,      sheet_name="Sessions",          index=False)
except PermissionError:
    writer_path = output_path.replace(".xlsx", "_LOCKED.xlsx")
    print(f"\n[WARNING] Permission denied: '{output_path}' appears to be open in Excel.")
    print(f"Saving to fallback path: '{writer_path}'")
    with pd.ExcelWriter(writer_path, engine="openpyxl") as writer:
        status_summary.to_excel(writer,   sheet_name="Status Summary",    index=False)
        employee_summary.to_excel(writer, sheet_name="Employee Summary",  index=False)
        sessions_df.to_excel(writer,      sheet_name="Sessions",          index=False)

# =====================================
# STYLING THE WORKBOOK (EXECUTIVE THEME)
# =====================================

print("Polishing and formatting Excel sheets...")
wb = load_workbook(writer_path)

# Fonts & Colors
font_family = "Calibri"  # Standard high-compatibility professional font
header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
title_font = Font(name=font_family, size=16, bold=True, color="1B4F72")
meta_font = Font(name=font_family, size=9, italic=True, color="5D6D7E")
data_font = Font(name=font_family, size=10, color="2C3E50")
bold_data_font = Font(name=font_family, size=10, bold=True, color="2C3E50")
kpi_label_font = Font(name=font_family, size=9, color="566573")
kpi_value_font = Font(name=font_family, size=14, bold=True, color="1B4F72")

header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")  # Deep slate/blue
zebra_fill = PatternFill(start_color="F8F9F9", end_color="F8F9F9", fill_type="solid")   # Very light gray
white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
kpi_fill = PatternFill(start_color="EAECEE", end_color="EAECEE", fill_type="solid")      # Medium-light gray

# Borders
thin_line = Side(border_style="thin", color="BDC3C7")
border_all = Border(left=thin_line, right=thin_line, top=thin_line, bottom=thin_line)
thick_bottom = Border(bottom=Side(border_style="medium", color="2C3E50"))
double_bottom = Border(bottom=Side(border_style="double", color="2C3E50"), top=Side(border_style="thin", color="BDC3C7"))

# Alignments
align_left = Alignment(horizontal="left", vertical="center")
align_right = Alignment(horizontal="right", vertical="center")
align_center = Alignment(horizontal="center", vertical="center")

# Determine reporting period dynamically
min_date = sessions_df["Date"].min().strftime("%d/%m/%Y")
max_date = sessions_df["Date"].max().strftime("%d/%m/%Y")
gen_time = datetime.now().strftime("%d/%m/%Y %H:%M")

# -----------------
# SHEET: Status Summary (Executive View)
# -----------------
ws_status = wb["Status Summary"]

# Insert 6 rows at the top for title block and KPIs
ws_status.insert_rows(1, 6)

# Title Block
ws_status.cell(row=1, column=1, value="OASIS SPECIALIST HOSPITAL").font = title_font
ws_status.cell(row=2, column=1, value=f"Reporting Period: {min_date} to {max_date}   |   Generated: {gen_time}").font = meta_font

# Organization KPIs calculation
total_employees = len(status_summary)
avg_late_pct = round(status_summary["Late Percentage"].mean(), 1)
total_shifts_logged = len(sessions_df)
incomplete_pct = round((status_summary["Missing C/In (Unknown Lateness)"].sum() / status_summary["Days Attended"].sum()) * 100, 1)

# Render KPI Cards in rows 4-5
def create_kpi_card(ws, col_start, label, val):
    ws.merge_cells(start_row=4, start_column=col_start, end_row=4, end_column=col_start+1)
    ws.merge_cells(start_row=5, start_column=col_start, end_row=5, end_column=col_start+1)
    
    cell_lbl = ws.cell(row=4, column=col_start, value=label)
    cell_lbl.font = kpi_label_font
    cell_lbl.alignment = align_center
    cell_lbl.fill = kpi_fill
    
    cell_val = ws.cell(row=5, column=col_start, value=val)
    cell_val.font = kpi_value_font
    cell_val.alignment = align_center
    cell_val.fill = kpi_fill
    
    for r in range(4, 6):
        for c in range(col_start, col_start+2):
            ws.cell(row=r, column=c).border = border_all

create_kpi_card(ws_status, 1, "EMPLOYEES AUDITED", total_employees)
create_kpi_card(ws_status, 3, "TOTAL SHIFTS LOGGED", total_shifts_logged)
create_kpi_card(ws_status, 5, "AVG LATE PERCENTAGE", f"{avg_late_pct}%")
create_kpi_card(ws_status, 7, "INCOMPLETE SHIFT RATE", f"{incomplete_pct}%")

# Style main table headers
header_row = 7
ws_status.row_dimensions[header_row].height = 28

for col in range(1, ws_status.max_column + 1):
    cell = ws_status.cell(row=header_row, column=col)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = align_center
    cell.border = border_all

# Format data rows
data_start_row = 8
max_row = ws_status.max_row

# Fill colors for conditional formatting
green_fill  = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")  # Pastel green
green_font  = Font(name=font_family, size=10, bold=True, color="385723")
amber_fill  = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # Pastel yellow
amber_font  = Font(name=font_family, size=10, bold=True, color="7F6000")
red_fill    = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")  # Pastel orange/red
red_font    = Font(name=font_family, size=10, bold=True, color="C65911")
gray_fill   = PatternFill(start_color="EAECEE", end_color="EAECEE", fill_type="solid")  # Pastel gray
gray_font   = Font(name=font_family, size=10, color="5D6D7E")

for row in range(data_start_row, max_row + 1):
    ws_status.row_dimensions[row].height = 20
    is_zebra = (row % 2 == 0)
    row_fill = zebra_fill if is_zebra else white_fill
    
    for col in range(1, ws_status.max_column + 1):
        cell = ws_status.cell(row=row, column=col)
        cell.font = data_font
        cell.fill = row_fill
        cell.border = border_all
        
        if col == 1:
            cell.alignment = align_left
        else:
            cell.alignment = align_center

# Resolve column numbers by header name (row 7 is the header row after insertion)
_hdr = {ws_status.cell(row=header_row, column=c).value: c
        for c in range(1, ws_status.max_column + 1)}
_late_pct_col = _hdr.get("Late Percentage")
_missing_cout_col = _hdr.get("Missing Checkout Days")

for row in range(data_start_row, max_row + 1):
    # Conditional formatting on Late Percentage
    if _late_pct_col:
        pct_cell = ws_status.cell(row=row, column=_late_pct_col)
        pct_cell.alignment = align_right
        pct_cell.font = bold_data_font
        val = pct_cell.value
        try:
            val_float = float(val)
            if val_float <= 10.0:
                pct_cell.fill = green_fill
                pct_cell.font = green_font
            elif val_float <= 30.0:
                pct_cell.fill = amber_fill
                pct_cell.font = amber_font
            else:
                pct_cell.fill = red_fill
                pct_cell.font = red_font
        except (ValueError, TypeError):
            pass

    # Highlight Missing Checkout Days
    if _missing_cout_col:
        inc_cell = ws_status.cell(row=row, column=_missing_cout_col)
        try:
            if int(inc_cell.value) > 0:
                inc_cell.fill = gray_fill
                inc_cell.font = gray_font
        except (ValueError, TypeError):
            pass

# Add Summary Row at the bottom
summary_row = max_row + 1
ws_status.cell(row=summary_row, column=1, value="Total / Average").font = bold_data_font
ws_status.cell(row=summary_row, column=1).alignment = align_left

_avg_cols = {_late_pct_col} if _late_pct_col else set()

for col in range(2, ws_status.max_column + 1):
    cell = ws_status.cell(row=summary_row, column=col)
    cell.font = bold_data_font
    cell.border = double_bottom
    col_letter = get_column_letter(col)
    if col in _avg_cols:
        cell.value = f"=AVERAGE({col_letter}{data_start_row}:{col_letter}{max_row})"
        cell.alignment = align_right
    else:
        cell.value = f"=SUM({col_letter}{data_start_row}:{col_letter}{max_row})"
        cell.alignment = align_center

# Freeze Top Row (row 7) and first column (Employee Name)
ws_status.freeze_panes = "B8"

# -----------------
# SHEET: Employee Summary
# -----------------
ws_emp = wb["Employee Summary"]
ws_emp.row_dimensions[1].height = 26

for col in range(1, ws_emp.max_column + 1):
    cell = ws_emp.cell(row=1, column=col)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = align_center
    cell.border = border_all

for row in range(2, ws_emp.max_row + 1):
    ws_emp.row_dimensions[row].height = 19
    is_zebra = (row % 2 == 0)
    row_fill = zebra_fill if is_zebra else white_fill
    for col in range(1, ws_emp.max_column + 1):
        cell = ws_emp.cell(row=row, column=col)
        cell.font = data_font
        cell.fill = row_fill
        cell.border = border_all
        if col == 1:
            cell.alignment = align_left
        elif col in [4, 5]:
            cell.alignment = align_right
            if cell.value is not None:
                cell.number_format = "0.00"
        else:
            cell.alignment = align_center

ws_emp.freeze_panes = "B2"

# -----------------
# SHEET: Sessions
# -----------------
ws_sess = wb["Sessions"]
ws_sess.row_dimensions[1].height = 26

status_col_idx = flag_col_idx = None
for col in range(1, ws_sess.max_column + 1):
    v = ws_sess.cell(1, col).value
    ws_sess.cell(1, col).font = header_font
    ws_sess.cell(1, col).fill = header_fill
    ws_sess.cell(1, col).alignment = align_center
    ws_sess.cell(1, col).border = border_all
    
    if v == "Attendance Status":
        status_col_idx = col
    if v == "Data Quality Flag":
        flag_col_idx = col

orange_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
orange_font = Font(name=font_family, size=10, color="7F6000")

for row in range(2, ws_sess.max_row + 1):
    ws_sess.row_dimensions[row].height = 19
    is_zebra = (row % 2 == 0)
    row_fill = zebra_fill if is_zebra else white_fill
    
    for col in range(1, ws_sess.max_column + 1):
        cell = ws_sess.cell(row=row, column=col)
        cell.font = data_font
        cell.fill = row_fill
        cell.border = border_all
        
        if col == 1:
            cell.alignment = align_left
        elif col in [4, 5]:  # Timestamps
            cell.alignment = align_center
            if cell.value is not None:
                cell.number_format = "yyyy-mm-dd hh:mm"
        elif col == 8:  # Hours worked
            cell.alignment = align_right
            if cell.value is not None:
                cell.number_format = "0.00"
        else:
            cell.alignment = align_center
            
    # Conditional format on Attendance Status
    if status_col_idx:
        stat_cell = ws_sess.cell(row=row, column=status_col_idx)
        val = stat_cell.value
        if val == "On Time":
            stat_cell.fill = green_fill
            stat_cell.font = green_font
        elif val == "Slightly Late":
            stat_cell.fill = amber_fill
            stat_cell.font = amber_font
        elif val in ["Late", "Very Late"]:
            stat_cell.fill = red_fill
            stat_cell.font = red_font
        else:
            stat_cell.fill = gray_fill
            stat_cell.font = gray_font
            
    # Highlight quality warnings
    if flag_col_idx:
        flag_cell = ws_sess.cell(row=row, column=flag_col_idx)
        if flag_cell.value and str(flag_cell.value).strip() != "":
            flag_cell.fill = orange_fill
            flag_cell.font = orange_font

ws_sess.freeze_panes = "B2"

# -----------------
# AUTO-FIT COLUMN WIDTHS FOR ALL SHEETS
# -----------------
for ws in wb.worksheets:
    for col_cells in ws.columns:
        if ws.title == "Status Summary" and col_cells[0].column in [1, 2, 3, 4, 5, 6, 7, 8] and any(cell.row in [4, 5] for cell in col_cells):
            widths = []
            for cell in col_cells:
                if cell.row > 6 and cell.value:
                    widths.append(len(str(cell.value)))
            width = max(widths) if widths else 10
        else:
            width = max((len(str(c.value)) if c.value else 0) for c in col_cells)
            
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = min(max(width + 4, 12), 45)

# =====================================
# SAVE WORKBOOK
# =====================================

wb.save(writer_path)
print("Attendance report generated successfully!")
print(f"Output File: {writer_path}")

print()
flagged = sessions_df[sessions_df["Data Quality Flag"] != ""]
print(f"Total session rows : {len(sessions_df)}")
print(f"Unique employee-days: {sessions_df[['Employee','Date']].drop_duplicates().shape[0]}")
print(f"Flagged records    : {len(flagged)}")
